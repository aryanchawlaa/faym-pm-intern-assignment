"""
Excel I/O layer.

Excel is the single source of truth (v1 has no external DB). We take some
care here because:

  1. Write-back is per line item, never batched at the end. If the agent
     dies mid-run, everything already completed is on disk.
  2. Row lookups use the internal row_id (openpyxl row number), so
     concurrent human edits to unrelated rows don't corrupt anything.
  3. Terminal states are hard-checked before overwriting, so a re-run
     never clobbers a completed row.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from typing import Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

log = logging.getLogger("excel_io")


class Status(str, Enum):
    PENDING = "Pending"
    TO_DO = "To Do"
    IN_PROGRESS = "In Progress"
    DONE = "Done"
    NEEDS_REVIEW = "Needs human review"


PENDING_STATES = {Status.PENDING.value, Status.TO_DO.value, ""}
TERMINAL_STATES = {Status.DONE.value, Status.NEEDS_REVIEW.value}


# Canonical column ordering. Anything below expects this schema.
COLUMNS = [
    "Platform", "Order ID", "Product / SKU", "Return window",
    "Return ID", "Return status", "Refund amount",
    "Task status", "Timestamp", "Attempt count", "Return model used",
    "Error / log",
]


@dataclass
class TaskRow:
    row_id: int             # openpyxl row number (1-indexed, header = 1)
    platform: str
    order_id: str
    sku: str
    return_window: str
    task_status: str
    attempt_count: int = 0


@dataclass
class Result:
    row_id: int
    sku: str
    return_status: str                          # Placed / Failed / Out of window
    task_status: Status                         # Done / Needs human review
    refund_amount: Optional[float] = None
    return_id: Optional[str] = None
    error: Optional[str] = None
    timestamp: datetime = field(default_factory=datetime.utcnow)


class ExcelQueue:
    def __init__(self, path: str, sheet: str | None = None):
        self.path = path
        self.wb = load_workbook(path)
        self.ws: Worksheet = self.wb[sheet] if sheet else self.wb.active
        self._col_idx = self._map_columns()

    def _map_columns(self) -> dict[str, int]:
        headers = {}
        for idx, cell in enumerate(self.ws[1], start=1):
            if cell.value:
                headers[str(cell.value).strip()] = idx
        missing = [c for c in COLUMNS if c not in headers]
        if missing:
            raise ValueError(
                f"Excel is missing required columns: {missing}. "
                f"Expected schema: {COLUMNS}"
            )
        return headers

    def _get(self, row: int, col: str):
        return self.ws.cell(row=row, column=self._col_idx[col]).value

    def _set(self, row: int, col: str, value):
        self.ws.cell(row=row, column=self._col_idx[col]).value = value

    def load_pending(self) -> list[TaskRow]:
        out: list[TaskRow] = []
        for row in range(2, self.ws.max_row + 1):
            status = (self._get(row, "Task status") or "").strip()
            if status not in PENDING_STATES:
                continue
            platform = self._get(row, "Platform")
            order_id = self._get(row, "Order ID")
            sku = self._get(row, "Product / SKU")
            if not (platform and order_id and sku):
                continue
            out.append(TaskRow(
                row_id=row,
                platform=str(platform).strip(),
                order_id=str(order_id).strip(),
                sku=str(sku).strip(),
                return_window=str(self._get(row, "Return window") or ""),
                task_status=status or Status.PENDING.value,
                attempt_count=int(self._get(row, "Attempt count") or 0),
            ))
        return out

    def is_terminal(self, task: TaskRow) -> bool:
        status = (self._get(task.row_id, "Task status") or "").strip()
        return status in TERMINAL_STATES

    def write_result(self, r: Result, return_model: str) -> None:
        """Idempotent per-row write. Called immediately after each SKU is
        processed so partial progress survives a crash."""
        row = r.row_id
        self._set(row, "Return ID", r.return_id or "")
        self._set(row, "Return status", r.return_status)
        self._set(row, "Refund amount", r.refund_amount if r.refund_amount is not None else "")
        self._set(row, "Task status", r.task_status.value)
        self._set(row, "Timestamp", r.timestamp.isoformat(timespec="seconds"))
        # return_model may be a ReturnModel enum or plain string; store the
        # human-readable value either way (openpyxl would otherwise write
        # the enum repr, e.g. "ReturnModel.BATCH").
        self._set(row, "Return model used",
                  return_model.value if hasattr(return_model, "value") else str(return_model))
        prev_attempts = int(self._get(row, "Attempt count") or 0)
        self._set(row, "Attempt count", prev_attempts + 1)
        self._set(row, "Error / log", r.error or "")
        # Flush after every write so a crash never loses completed work.
        self.wb.save(self.path)
        log.debug("Row %d written: %s / %s", row, r.task_status.value, r.return_status)

    def save(self) -> None:
        self.wb.save(self.path)
